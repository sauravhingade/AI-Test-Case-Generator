import json
import pandas as pd
from io import BytesIO
from backend.schemas.models import TestSuite, TestCase


# ── Flatten TestCase to dict ────────────────────────────────


def _flatten_test_case(tc: TestCase, suite_name: str) -> dict:
    """
    Flattens a TestCase into a flat dict for CSV/Excel export.
    BDD steps are joined as readable text.
    """
    steps_text = "\n".join(f"{step.keyword} {step.action}" for step in tc.steps)

    return {
        "TC ID": tc.id,
        "Suite": suite_name,
        "User Story ID": tc.user_story_id,
        "Title": tc.title,
        "Test Type": tc.test_type.value,
        "Priority": tc.priority.value,
        "Is Edge Case": tc.is_edge_case,
        "Preconditions": " | ".join(tc.preconditions),
        "BDD Steps": steps_text,
        "Expected Result": tc.expected_result,
        "Tags": ", ".join(tc.tags),
    }


def _suites_to_dataframe(suites: list[TestSuite]) -> pd.DataFrame:
    """Converts all TestSuites into a single flat DataFrame."""
    rows = []
    for suite in suites:
        for tc in suite.test_cases:
            rows.append(_flatten_test_case(tc, suite.suite_name))
    return pd.DataFrame(rows)


# ── Export functions ────────────────────────────────────────


def export_as_json(suites: list[TestSuite]) -> bytes:
    """Exports test suites as formatted JSON bytes."""
    data = {
        "total_test_cases": sum(s.total_count for s in suites),
        "total_suites": len(suites),
        "suites": [suite.model_dump() for suite in suites],
    }
    return json.dumps(data, indent=2, default=str).encode("utf-8")


def export_as_csv(suites: list[TestSuite]) -> bytes:
    """Exports test suites as CSV bytes. One row per test case."""
    df = _suites_to_dataframe(suites)
    buffer = BytesIO()
    df.to_csv(buffer, index=False, encoding="utf-8")
    return buffer.getvalue()


def export_as_excel(suites: list[TestSuite]) -> bytes:
    """
    Exports as Excel bytes.
    Sheet 1: All Test Cases (flat, color coded by priority)
    Sheet 2: Summary (suite-level stats)
    """
    df_cases = _suites_to_dataframe(suites)

    summary_rows = []
    for suite in suites:
        summary_rows.append(
            {
                "Suite Name": suite.suite_name,
                "User Story ID": suite.user_story_id,
                "Total Cases": suite.total_count,
                "Critical": suite.critical_count,
                "Edge Cases": suite.edge_case_count,
                "Coverage Areas": ", ".join(suite.coverage_areas),
            }
        )
    df_summary = pd.DataFrame(summary_rows)

    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df_cases.to_excel(writer, sheet_name="Test Cases", index=False)
        df_summary.to_excel(writer, sheet_name="Summary", index=False)

        # ── Color code priority column ──
        from openpyxl.styles import PatternFill, Font

        ws = writer.sheets["Test Cases"]

        priority_colors = {
            "critical": ("FF4C4C", "FFFFFF"),
            "high": ("FF9900", "FFFFFF"),
            "medium": ("FFD700", "000000"),
            "low": ("90EE90", "000000"),
        }

        priority_col = None
        for col_idx, cell in enumerate(ws[1], start=1):
            if cell.value == "Priority":
                priority_col = col_idx
                break

        if priority_col:
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                cell = row[priority_col - 1]
                priority_val = str(cell.value).lower() if cell.value else ""
                if priority_val in priority_colors:
                    bg, fg = priority_colors[priority_val]
                    cell.fill = PatternFill(
                        start_color=bg, end_color=bg, fill_type="solid"
                    )
                    cell.font = Font(color=fg, bold=True)

        # ── Style header row ──
        from openpyxl.styles import Alignment

        header_fill = PatternFill(
            start_color="2D2D2D", end_color="2D2D2D", fill_type="solid"
        )
        header_font = Font(color="FFFFFF", bold=True)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )
        ws.row_dimensions[1].height = 30

        # ── Style header row in Summary sheet ──
        ws2 = writer.sheets["Summary"]
        for cell in ws2[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws2.row_dimensions[1].height = 30

        # Auto-fit column widths
        for sheet in writer.sheets.values():
            for col in sheet.columns:
                max_len = max(
                    (len(str(cell.value)) for cell in col if cell.value), default=10
                )
                sheet.column_dimensions[col[0].column_letter].width = min(
                    max_len + 4, 60
                )

    return buffer.getvalue()
